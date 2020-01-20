#!/usr/bin/env python
# -*- coding: utf-8 -*-

import random
import numpy as np
import pandas as pd
from sklearn.svm import SVC
from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
from sklearn.cross_validation import StratifiedKFold
from sklearn import metrics
from sklearn.metrics import roc_curve, roc_auc_score, auc
import pickle
from sklearn.metrics import make_scorer
from sklearn.metrics import average_precision_score,precision_recall_curve
from sklearn.model_selection import cross_val_score,cross_validate
import openpyxl
import math

def score_auc(X, y):
	fpr, tpr, threshold = roc_curve(X, y) 
	roc = roc_auc_score(X, y)
	wb = openpyxl.load_workbook(filename=filename)
	ws1 = wb['Sheet1']
	ws2 = wb['Sheet2']
	col1 = ws1.max_column +1
	col2 = ws2.max_column +1
	for row, entry in enumerate(fpr, start=1):
		ws1.cell(row=row, column=col1, value=entry)
	for row, entry in enumerate(tpr, start=1):
		ws2.cell(row=row, column=col2, value=entry)
	wb.save(filename)
	return roc

def score_ap(X, y):
	precision, recall, threshold = precision_recall_curve(X, y)
	avg_pre = average_precision_score(X,y)
	wb = openpyxl.load_workbook(filename=filename1)
	ws1 = wb['Sheet1']
	ws2 = wb['Sheet2']
	col1 = ws1.max_column +1
	col2 = ws2.max_column +1
	for row, entry in enumerate(precision, start=1):
		ws1.cell(row=row, column=col1, value=entry)
	for row, entry in enumerate(recall, start=1):
		ws2.cell(row=row, column=col2, value=entry)
	wb.save(filename1)
	return avg_pre

def model_training(X_res,y_res,mo):
	custom_auc = make_scorer(score_auc, needs_threshold = True)
	custom_ap = make_scorer(score_ap, needs_threshold = True)
	#print("Function  Called")
	scoring = {'acc': 'accuracy','recall': 'recall','pre': 'precision','ap': 'average_precision','roc_auc': 'roc_auc','f1':'f1','f1_micro':'f1_micro', 'f1_macro':'f1_macro','custom_auc':custom_auc, 'custom_ap':custom_ap}
	scores = cross_validate(mo, X_res, y_res, scoring=scoring, cv=StratifiedKFold(y = y_res, n_folds = 5),return_train_score=False) 
	#print("\nReturned")
	return scores

if __name__ == "__main__":
	classifier1 = SVC(probability = True)					# SVM Classification
	classifier2 = LogisticRegression()						# LR Classification
	classifier3 = MLPClassifier()							# MLP Classification
	M = [classifier1, classifier2, classifier3]
	cls = ["SVM", "LR", "MLP"]		
	X  = pd.read_excel('C:\\Users\\Ajay Rastogi\\Desktop\\Extended Version (Journal)\\PeerJ\\PeerJ-Paper\\Data\\YelpZip\\Resulting Features\\Reviewer_Centric_Behavioral_Features.xlsx', sheetname=0)			# Change the path according to the file location
	X['MNR(u)'] =(X['MNR(u)']-X['MNR(u)'].min())/(X['MNR(u)'].max()-X['MNR(u)'].min())										# Min-max normalization on MNR
	X1 = pd.read_excel('C:\\Users\\Ajay Rastogi\\Desktop\\Extended Version (Journal)\\PeerJ\\PeerJ-Paper\\Data\\YelpZip\\Resulting Features\\Reviewer_Centric_Textual_Features.xlsx', sheetname=0)			# Change the path according to the file location
	X1['ARL(u)'] =(X1['ARL(u)']-X1['ARL(u)'].min())/(X1['ARL(u)'].max()-X1['ARL(u)'].min())									# Min-max normalization on ARL
	#print(X.head())
	#print(X.shape)
	#print(X1.head())
	#print(X1.shape)
	Y= pd.read_excel('C:\\Users\\Ajay Rastogi\\Desktop\\Extended Version (Journal)\\PeerJ\\PeerJ-Paper\\Data\\YelpZip\\Labels_for_reviewers.xlsx', sheetname=0)			# Change the path according to the file location
	Y = Y.iloc[:, 0:1]
	Y.ix[(Y.Label == 1), 'Label'] = 0
	Y.ix[(Y.Label == -1), 'Label'] = 1
	#print(Y.head())
	#print(Y.shape)
	
	# Code for Balancing the Feature Set ........................................
	X_new = pd.concat([X,X1,Y],axis =1)
	X_new_spam = X_new.loc[X_new['Label'] == 1]				# Total spammers
	X_new_notspam = X_new.loc[X_new['Label'] == 0]			# Total non-spammers
	X_new_spam = X_new_spam.iloc[np.random.permutation(len(X_new_spam))]				# Shuffle spam
	X_new_notspam = X_new_notspam.iloc[np.random.permutation(len(X_new_notspam))]		# Shuffle not-spam
	partitions = int(math.ceil(X_new_notspam.shape[0]/float(X_new_spam.shape[0])))
	X_df = np.array_split(X_new_notspam, partitions)		# split majority class samples(not-spam class) into nearly equal size partitions 
	X_bal_sets = []
	for i in range(0, partitions):
		X_bal_sets.append(pd.concat([X_new_spam,X_df[i]],axis=0))	# Merging partitions with spam-class to get balanced instance sets 
	
	print("First partition Distribution : " , X_bal_sets[0]['Label'].value_counts())
	print("Last partition Distribution : " , X_bal_sets[partitions-1]['Label'].value_counts())
	# code end....................................................................


	"""for i in range(0,3):
		print("Model training on "+cls[i]+" classifier running...")
		filename = 'Data\\Results\\YelpZip\\Reviewer Hybrid\\'+cls[i]+'\\FPR_TPR_Threshold_Reviewer_Hybrid.xlsx'		# File for Writing FPR and TPR values at different thresholds for ROC curve analysis
		filename1 = 'Data\\Results\\YelpZip\\Reviewer Hybrid\\'+cls[i]+'\\Pre_Recall_Threshold_Reviewer_Hybrid.xlsx'	# File for Writing Precision and Recall values at different thresholds for Precision-recall curve analysis
		f = 'Data\\Results\\YelpZip\\Reviewer Hybrid\\'+cls[i]+'\\X_bal_sets.sav'
		pickle.dump(X_bal_sets, open(f, 'wb'))				# Saving balanced instance sets as an object 
		arg = []
		for u in range(len(X_bal_sets)):
			arg.append([X_bal_sets[u].iloc[:,0:X_bal_sets[u].shape[1]-1], X_bal_sets[u]['Label'],M[i]])
		result = []
		ss = 1
		for sample in arg:
			print("Model training on Partition : "+str(ss)+"\n")
			result.append(model_training(sample[0],sample[1],sample[2]))
			ss = ss + 1
		k = 1
		for mm in result:
			filename2 = 'Data\\Results\\YelpZip\\Reviewer Hybrid\\'+cls[i]+'\\Model_'+str(k)+'.sav'				# Change the path according to the file location
			pickle.dump(mm, open(filename2, 'wb'))
			k = k+1
"""
